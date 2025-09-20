from pathlib import Path

from openpyxl import Workbook

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

workbook = Workbook()

# Create sheet
sheet_name = "Sheet Name"
workbook.create_sheet(sheet_name)
worksheet = workbook[sheet_name]
workbook.remove(workbook["Sheet"])

# Insert headers
worksheet.cell(1, 1, "Name")
worksheet.cell(1, 2, "Age")
worksheet.cell(1, 3, "From")

# Insert data - Option 1
students = [
    ["Maria", 20, "Argentina"],
    ["John", 40, "United States"],
    ["Julia", 20, "Brazil"],
]

for i, student_row in enumerate(students, start=2):
    for j, student_column in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_column)

# Insert data - Option 2
students = [
    ["Carol", 30, "Portugal"],
    ["Oliver", 10, "England"],
]

for student in students:
    worksheet.append(student)

# Update cell
worksheet["B3"].value = 18

# Read data
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end="\t")

        # Find and update cell
        if cell.value == "Oliver":
            worksheet.cell(cell.row, 3, "Ireland")

    print()

workbook.save(WORKBOOK_PATH)
